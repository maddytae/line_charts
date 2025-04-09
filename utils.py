import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

def add_table_of_contents(excel_path):
    # Load the workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Get all existing sheet names
    original_sheets = wb.sheetnames
    
    # Create new Table_of_Contents sheet
    toc_sheet = wb.create_sheet("Table_of_Contents", 0)  # 0 means insert at beginning
    
    # Add header
    toc_sheet['A1'] = "Table of Contents"
    toc_sheet['A1'].font = openpyxl.styles.Font(bold=True)
    
    # Add sheet names with hyperlinks
    for i, sheet_name in enumerate(original_sheets, start=2):
        cell = f'A{i}'
        toc_sheet[cell] = sheet_name
        
        # Create hyperlink to the respective sheet
        hyperlink = Hyperlink(ref=f"{sheet_name}!A1", 
                            location=f"'{sheet_name}'!A1",
                            tooltip=f"Go to {sheet_name}",
                            display=sheet_name)
        toc_sheet[cell].hyperlink = hyperlink
        toc_sheet[cell].style = "Hyperlink"
    
    # Add return hyperlink to each original sheet
    for sheet_name in original_sheets:
        sheet = wb[sheet_name]
        # Create hyperlink back to Table_of_Contents
        hyperlink = Hyperlink(ref="Table_of_Contents!A1",
                            location="'Table_of_Contents'!A1",
                            tooltip="Return to Table of Contents",
                            display="Table of Contents")
        sheet['A1'].hyperlink = hyperlink
        sheet['A1'].style = "Hyperlink"
    
    # Adjust column width for better visibility
    toc_sheet.column_dimensions['A'].width = 20
    
    # Save the modified workbook
    wb.save(excel_path)
    wb.close()

# Example usage:
# add_table_of_contents("path/to/your/excel_file.xlsx")
